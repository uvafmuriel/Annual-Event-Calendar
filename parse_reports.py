"""
CGZ Board — Weekly XLSX Parser
Reads up to 4 weekly report files from /weekly-reports/,
extracts the upcoming Monday–Friday events,
and merges them into cgz-data.json preserving
announcements, alerts, work orders, and lunch data.
"""

import json
import os
import glob
from datetime import datetime, timedelta, date
import openpyxl

REPORTS_DIR = "weekly-reports"
OUTPUT_FILE = "cgz-data.json"

# Skip these event types — they're internal holds, not displayable events
SKIP_TYPES = {"Maintenance"}
SKIP_NAME_PREFIXES = ("maintenance hold", "sa event mgt hold")

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_ABBRS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
DAY_COLORS = {
    "Monday":    "#2980b9",
    "Tuesday":   "#1a7a40",
    "Wednesday": "#6b2fa0",
    "Thursday":  "#c25e00",
    "Friday":    "#c0392b",
}


def get_next_week_range():
    """Get the upcoming Mon–Fri date range.
    If today is Friday or weekend, get next Mon–Fri.
    Otherwise get the current week's Mon–Fri.
    """
    today = date.today()
    weekday = today.weekday()  # 0=Mon, 6=Sun

    # If Friday (4), Saturday (5), Sunday (6) → get NEXT Monday
    if weekday >= 4:
        days_to_monday = 7 - weekday
    else:
        # Mon–Thu → get THIS week's Monday
        days_to_monday = -weekday

    monday = today + timedelta(days=days_to_monday)
    friday = monday + timedelta(days=4)
    return monday, friday


def format_time(dt):
    """Format datetime to 12-hour time string."""
    if dt is None:
        return ""
    if hasattr(dt, 'strftime'):
        return dt.strftime("%-I:%M %p").replace("AM", "am").replace("PM", "pm")
    return str(dt)


def clean_name(name):
    """Truncate long event names cleanly."""
    if not name:
        return ""
    name = str(name).strip()
    # Truncate at 45 chars with ellipsis
    if len(name) > 45:
        name = name[:42].rstrip() + "…"
    return name


def parse_xlsx(filepath):
    """Parse a single XLSX report file. Returns list of event dicts."""
    events = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if "Reservations" not in wb.sheetnames:
            print(f"  ⚠ No 'Reservations' sheet in {filepath}")
            return events

        ws = wb["Reservations"]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return events

        # Find header row (first row with 'Event Name')
        header_row = None
        data_start = 0
        for i, row in enumerate(rows):
            if row and row[0] == "Event Name":
                header_row = row
                data_start = i + 1
                break

        if header_row is None:
            print(f"  ⚠ Could not find header row in {filepath}")
            return events

        # Map column names to indices
        col = {str(h).strip(): i for i, h in enumerate(header_row) if h}

        for row in rows[data_start:]:
            if not row or not row[col.get("Event Name", 0)]:
                continue

            name       = row[col.get("Event Name", 0)]
            day        = row[col.get("Day", 1)]
            start      = row[col.get("Event Start", 4)]
            end        = row[col.get("Event End", 5)]
            location   = row[col.get("Location", 8)]
            event_type = row[col.get("Event Type", 17)]
            state      = row[col.get("Event State", 18)]

            # Skip maintenance and internal holds
            if event_type in SKIP_TYPES:
                continue
            if str(name).strip().lower().startswith(SKIP_NAME_PREFIXES):
                continue

            # Only show Confirmed and Tentative
            if state not in ("Confirmed", "Tentative"):
                continue

            # Must have a valid date
            if not day or not hasattr(day, 'date'):
                continue

            event_date = day.date() if hasattr(day, 'date') else day

            events.append({
                "date":     event_date,
                "name":     clean_name(name),
                "time":     format_time(start),
                "end_time": format_time(end),
                "location": str(location).strip() if location else "",
                "type":     str(event_type).strip() if event_type else "",
                "state":    str(state).strip() if state else "",
            })

        wb.close()
    except Exception as e:
        print(f"  ✗ Error parsing {filepath}: {e}")

    return events


def build_week_of_label(monday):
    """Build a readable week label like 'June 2 – 6, 2025'."""
    friday = monday + timedelta(days=4)
    if monday.month == friday.month:
        return f"{monday.strftime('%B %-d')} – {friday.strftime('%-d, %Y')}"
    else:
        return f"{monday.strftime('%B %-d')} – {friday.strftime('%B %-d, %Y')}"


def main():
    print("CGZ Board Parser — starting")

    # Find all XLSX files in weekly-reports/
    xlsx_files = glob.glob(os.path.join(REPORTS_DIR, "*.xlsx"))
    if not xlsx_files:
        print(f"  No XLSX files found in {REPORTS_DIR}/")
        print("  Skipping update — cgz-data.json unchanged")
        return

    print(f"  Found {len(xlsx_files)} file(s): {[os.path.basename(f) for f in xlsx_files]}")

    # Parse all files, collect all events
    all_events = []
    for f in xlsx_files:
        print(f"  Parsing {os.path.basename(f)}…")
        evs = parse_xlsx(f)
        print(f"    → {len(evs)} events")
        all_events.extend(evs)

    # Get the upcoming Mon–Fri
    monday, friday = get_next_week_range()
    print(f"\n  Target week: {monday} → {friday}")

    # Filter to just that week (Mon–Fri)
    week_events = [e for e in all_events if monday <= e["date"] <= friday]
    print(f"  Events this week: {len(week_events)}")

    # Group by weekday name
    days_map = {name: [] for name in ["Monday","Tuesday","Wednesday","Thursday","Friday"]}
    for ev in sorted(week_events, key=lambda e: (e["date"], e["time"])):
        day_name = DAY_NAMES[ev["date"].weekday()]
        if day_name in days_map:
            detail = ev["location"]
            if ev["state"] == "Tentative":
                detail += " (Tentative)" if detail else "Tentative"
            days_map[day_name].append({
                "name":   ev["name"],
                "time":   ev["time"],
                "detail": detail,
            })

    # Load existing cgz-data.json to preserve announcements, alerts, etc.
    existing = {}
    if os.path.exists(OUTPUT_FILE):
        try:
            with open(OUTPUT_FILE) as f:
                existing = json.load(f)
            print(f"  Loaded existing {OUTPUT_FILE}")
        except Exception as e:
            print(f"  ⚠ Could not load existing data: {e}")

    # Build updated days array — keep colors from existing if present
    existing_days = {d["name"]: d for d in existing.get("days", [])}
    days_array = []
    for name in ["Monday","Tuesday","Wednesday","Thursday","Friday"]:
        existing_day = existing_days.get(name, {})
        days_array.append({
            "name":   name,
            "abbr":   name[:3].upper(),
            "color":  existing_day.get("color", DAY_COLORS[name]),
            "events": days_map[name],
        })

    # Build final data — preserve everything except days and weekOf
    output = {
        **existing,
        "weekOf": build_week_of_label(monday),
        "days":   days_array,
    }

    # Remove weather — it's pulled live from Open-Meteo
    output.pop("weather", None)

    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\n  ✓ Written to {OUTPUT_FILE}")
    print(f"  Week of: {output['weekOf']}")
    for day in days_array:
        print(f"  {day['name']}: {len(day['events'])} event(s)")
        for ev in day['events']:
            print(f"    {ev['time']} — {ev['name']} @ {ev['detail']}")


if __name__ == "__main__":
    main()
