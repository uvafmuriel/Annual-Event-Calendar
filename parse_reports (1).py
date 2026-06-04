"""
CGZ Board — Weekly XLSX Parser
Reads XLSX files from weekly-reports/,
parses events for the upcoming Mon-Fri,
and writes cgz-data.json preserving
announcements, alerts, work orders, and lunch.
"""

import json, os, glob
from datetime import date, timedelta
import openpyxl

REPORTS_DIR = "weekly-reports"
OUTPUT_FILE = "cgz-data.json"

SKIP_TYPES       = {"Maintenance"}
SKIP_NAME_STARTS = ("maintenance hold", "sa event mgt hold")
DAY_NAMES        = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
DAY_COLORS       = {
    "Monday":    "#2980b9",
    "Tuesday":   "#1a7a40",
    "Wednesday": "#6b2fa0",
    "Thursday":  "#c25e00",
    "Friday":    "#c0392b",
}

def get_week_range():
    today = date.today()
    wd = today.weekday()
    days_to_monday = (7 - wd) if wd >= 4 else -wd
    monday = today + timedelta(days=days_to_monday)
    return monday, monday + timedelta(days=4)

def week_label(monday):
    friday = monday + timedelta(days=4)
    if monday.month == friday.month:
        return f"{monday.strftime('%B %-d')} \u2013 {friday.strftime('%-d, %Y')}"
    return f"{monday.strftime('%B %-d')} \u2013 {friday.strftime('%B %-d, %Y')}"

def fmt_time(dt):
    if not dt or not hasattr(dt, 'strftime'): return ""
    return dt.strftime("%-I:%M %p")

def clean_name(name):
    name = str(name or "").strip()
    return name[:42].rstrip() + "\u2026" if len(name) > 45 else name

def parse_xlsx(filepath):
    events = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if "Reservations" not in wb.sheetnames:
            print(f"  No Reservations sheet in {filepath}")
            return events
        ws = wb["Reservations"]
        rows = list(ws.iter_rows(values_only=True))
        header_row, data_start = None, 0
        for i, row in enumerate(rows):
            if row and row[0] == "Event Name":
                header_row, data_start = row, i + 1
                break
        if not header_row: return events
        col = {str(h).strip(): i for i, h in enumerate(header_row) if h}
        for row in rows[data_start:]:
            if not row or not row[col.get("Event Name", 0)]: continue
            name       = row[col.get("Event Name", 0)]
            day        = row[col.get("Day", 1)]
            start      = row[col.get("Event Start", 4)]
            end        = row[col.get("Event End", 5)]
            location   = row[col.get("Location", 8)]
            event_type = row[col.get("Event Type", 17)]
            state      = row[col.get("Event State", 18)]
            if event_type in SKIP_TYPES: continue
            if str(name).strip().lower().startswith(SKIP_NAME_STARTS): continue
            if state not in ("Confirmed", "Tentative"): continue
            if not day or not hasattr(day, 'date'): continue
            events.append({
                "date":     day.date(),
                "name":     clean_name(name),
                "time":     fmt_time(start),
                "end_time": fmt_time(end),
                "location": str(location or "").strip(),
                "type":     str(event_type or "").strip(),
                "state":    str(state or "").strip(),
            })
        wb.close()
    except Exception as e:
        print(f"  Error parsing {filepath}: {e}")
    return events

def main():
    print("CGZ Board Parser")
    xlsx_files = glob.glob(os.path.join(REPORTS_DIR, "*.xlsx"))
    if not xlsx_files:
        print(f"No XLSX files in {REPORTS_DIR}/ -- nothing to do")
        return

    print(f"Found {len(xlsx_files)} file(s)")
    all_events = []
    for f in xlsx_files:
        print(f"Parsing {os.path.basename(f)}...")
        evs = parse_xlsx(f)
        print(f"  -> {len(evs)} events")
        all_events.extend(evs)

    monday, friday = get_week_range()
    wlabel = week_label(monday)
    print(f"Target week: {monday} -> {friday} ({wlabel})")

    week_events = sorted(
        [e for e in all_events if monday <= e["date"] <= friday],
        key=lambda e: (e["date"], e["time"])
    )
    print(f"Events this week: {len(week_events)}")

    # Load existing data to preserve announcements, alerts, work orders, lunch
    existing = {}
    if os.path.exists(OUTPUT_FILE):
        try:
            with open(OUTPUT_FILE) as f:
                existing = json.load(f)
            print(f"Loaded existing {OUTPUT_FILE}")
        except Exception as e:
            print(f"Could not load existing data: {e}")

    # Build days array
    existing_days = {d["name"]: d for d in existing.get("days", [])}
    days_map = {n: [] for n in ["Monday","Tuesday","Wednesday","Thursday","Friday"]}
    for ev in week_events:
        day_name = DAY_NAMES[ev["date"].weekday()]
        if day_name in days_map:
            detail = ev["location"]
            if ev["state"] == "Tentative":
                detail += " (Tentative)" if detail else "Tentative"
            days_map[day_name].append({
                "name":   ev["name"],
                "time":   ev["time"],
                "detail": detail,
            })

    days_array = []
    for name in ["Monday","Tuesday","Wednesday","Thursday","Friday"]:
        existing_day = existing_days.get(name, {})
        days_array.append({
            "name":   name,
            "abbr":   name[:3].upper(),
            "color":  existing_day.get("color", DAY_COLORS[name]),
            "events": days_map[name],
        })

    output = {
        **existing,
        "weekOf": wlabel,
        "days":   days_array,
    }
    output.pop("weather", None)

    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Done! Written to {OUTPUT_FILE}")
    for day in days_array:
        print(f"  {day['name']}: {len(day['events'])} event(s)")

if __name__ == "__main__":
    main()
